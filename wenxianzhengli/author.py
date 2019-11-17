# -*- coding: utf-8 -*-
"""
Created on Wed Apr 17 14:18:46 2019

@author: wyr
"""

from openpyxl import load_workbook
import os.path
import re

title = 'Removed Due to Privacy Concerns'
folder_spliter = '\\'
pdf_folder = 'papers\\'
xlsx_folder = 'xlsx\\'

class FileFinder():
    def __init__(self):
        self.flist = []
    
    def add_file_list(self, file_dir, suffix=''):
        l = len(suffix)
        if (l>0):
            for root_path, dirs, files in os.walk(file_dir):
                for f in files:
                    if (f[-1*l:] == suffix):
                        self.flist.append(os.path.join(root_path, f))
        else:
             for root_path, dirs, files in os.walk(file_dir):
                for f in files:
                    self.flist.append(os.path.join(root_path, f))           
    
    def add_file(self, f):
        self.flist.append(f)
        
    def get_list(self):
        return self.flist[:]
    
    def clear(self):
        self.flist = []

class AuthorLoader():
    def __init__(self):
        self.buffer = []
        
    def load_author(self, title, title_colown, flist, clip=10):
        title = re.sub(r'[^\w]', '', title).lower()
        for f in flist:
            filename = f.split(folder_spliter)[-1]
            print('loading ' + filename)
            wb = load_workbook(f)
            for s in wb.sheetnames:
                sheet = wb[s]
                for i in range(1, sheet.max_row+1):
                    line = [j.value for j in sheet[str(i)]]
                    try:
                        t = re.sub(r'[^\w]', '', line[title_colown]).lower()
                    except Exception as e:
                        print(e)
                        print('line ' + str(i))
                        print(line[title_colown])
                        t = ''
                    if (t == title):
                        self.buffer.append(line[:clip])
            wb.close()
    
    def get_author(self):
        return self.buffer[:]
    
    def clear(self):
        self.buffer = []

class AuthorMatcher():
    def __init__(self):
        self.buffer = [[],[],[]]
        
    def match_author(self, title_list, title_colown, author_list, name_colomn, country_colomn):
        for i in range(0, len(author_list)):
            author_list[i][title_colown] = re.sub(r'[^\w]', '', author_list[i][title_colown])
            author_list[i][title_colown] = author_list[i][title_colown].lower()
        for t in title_list:
            print('matching ' + title)
            self.buffer[0].append(t)
            t = re.sub(r'[^\w]', '', t).lower()
            author = []
            country = []
            for i in range(0, len(author_list)):
                if (author_list[i][title_colown] == t):
                    author.append(author_list[i][name_colomn])
                    country.append(author_list[i][country_colomn])
            author = list(set(author))
            country = list(set(country))     
            self.buffer[1].append(str(author))
            self.buffer[2].append(str(country))       
    
    def get_author(self):
        return self.buffer[:]
    
    def clear(self):
        self.buffer = [[],[],[]]

if __name__ == '__main__':
    flist_generator = FileFinder();
    flist_generator.add_file_list(xlsx_folder)
    author_loader = AuthorLoader()
    author_loader.load_author(title, 0, flist_generator.get_list())
    flist_generator.clear();
    
    flist_generator.add_file_list(pdf_folder)
    title_list = flist_generator.get_list()
    for i in range(0, len(title_list)):
        title_list[i] = title_list[i].split(folder_spliter)[-1][:-4]
    author_matcher = AuthorMatcher()
    author_matcher.match_author(title_list, 1, author_loader.get_author(), 5, 6)
    
    out = author_matcher.get_author()
    for i in range(0, len(out[0])):
        print(out[0][i])
    for i in range(0, len(out[0])):
        print(str(out[1][i])[1:-1])
    for i in range(0, len(out[0])):
        print(str(out[2][i])[1:-1])